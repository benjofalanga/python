from cursesmenu import SelectionMenu
import openpyxl
import sys
from openpyxl.utils import get_column_letter
import os
# used for filtering through exec .xlsx files only

def get_columns(sheet,list):
    print ('you can search by:')
    c=0
    le=['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P']
    for i in le:

        if (sheet[i+'1'].value==None):
            break
        else:
            # print(sheet[i+'1'].value)
            c+=1
            list.append(sheet[i+'1'].value)
    print ('Number of columns is ', c)
def chk_val(value,list):
    if value in list:
        print (value+" is in list")
    else:
        while value not in list:
            print ('wrong value please type again or type "q" to quit')
            value=input()
            if value=="q":
                sys.exit()
    return list

def get_col(sheet,value):
    for i in sheet [1]:
        if i.value==value:
            num0=i
    return num0

def filt(sheet,value,colm):
    print('input info to filter by:')
    info=input()
    print ("these are the " +info+ '\'' +'s emails:')
    users=[]
    k=0
    num=get_col(sheet,value)
    num=get_column_letter(num.column)
    for i in sheet[num]:
        if i.value== info:
            k+=1
            users.append(sheet[colm+str(i.row)].value)
            # print(sheet[colm+str(i.row)].value)
    print('there are '+str(k)+" " +info)
    return users
def filt0(sheet,value):
    users=[]
    k=0
    num1=get_col(sheet,value)
    num=get_column_letter(num1.column)
    for i in sheet[num]:
        if sheet[num+str(i.row)].value != None:
            users.append(sheet[num+str(i.row)].value)
    return users
def lowerc(list):
    nelist=[]
    for i in list:
        k=i.lower()
        nelist.append(k)
    return nelist
def menu(a_list):
    menu = SelectionMenu(a_list,"Select an excel file")
    menu.show()
    menu.join()
    selection = menu.selected_option
    return a_list[selection]

def ifexc(path):
    print('')
# def curdir(path):
#     q=0
#     if path in os.listdir():
#           q=1
#     else:
#         print("File not in current dir")
#     return q
#
# q=0
# while q==0:
print ('Select the name of the 1-st excel file (press enter for menu)')
input()

q=os.listdir()
file=menu(q)
print('you have selected '+file)

wb=openpyxl.load_workbook(file)
# wb=openpyxl.load_workbook('Adobe users.xlsx')
names0=wb.sheetnames[0]


sheet=wb[names0]
list=[]
print ('Select the name of the 2-nd excel file (press enter for menu)')
input()
file1=menu(q)
print('you have selected '+file)
wb1=openpyxl.load_workbook(file1)
# wb1=openpyxl.load_workbook('TammosExport_GMAIL.xlsx')
names1=wb1.sheetnames[0]
sheet1=wb1[names1]
list1=[]
get_columns(sheet,list)
get_columns(sheet1,list1)
print('The columns to search by in first file are :')
for i in list:
    print (i)
print('The columns to search by in second file are : ')
for i in list1:
    print (i)

print ("input which atribute to search by in 1-st file? (it's case sensitive!!!)")
value=input()

chk_val(value,list)
an=''
while an!='y' and an!='n':

    print('Do you want to filter additionally?/type y or n')
    an=input()

    if an=='y':
        print ('which column to get users from?')
        colm=input()
        comp1=filt(sheet,value,colm)
    elif an=='n':
        comp1=filt0(sheet,value)
        print ('continue with second file')
    else:

        print ('please type y or n or ctr+c to quit program')
print ("input which atribute to search by or (it's case sensitive!!!)")
value=input()
chk_val(value,list1)
an=''
while an!='y' and an!='n':

    print('Do you want to filter additionally?/type y or n')
    an=input()

    if an=='y':
        print ('which column to get users from?')
        colm2=input()
        comp2=filt(sheet1,value,colm2)
    elif an=='n':
        comp2=filt0(sheet1,value)
    else:

        print ('please type y or n or ctr+c to quit program')
diff=[]
comp1=lowerc(comp1)
comp2=lowerc(comp2)
for i in comp1:
    if i not in comp2:
        diff.append(i)
print('working...')
print('the different cells values are :')
for i in diff:
    print (i)
